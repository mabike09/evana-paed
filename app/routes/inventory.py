# app/routes/inventory.py
from decimal import Decimal
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, send_file
from flask_login import login_required, current_user
from sqlalchemy import func, or_, and_
from ..permissions import roles_required
from ..extensions import db
from ..forms import StockTxnForm, ItemPriceForm, VisitConsumeForm, BulkStockTxnForm
from ..models import (
    Item, ItemTxn, DispenseTxn, Visit, Invoice, InvoiceLine, Patient,
    PriceBook, PriceItem, Payer
)


def _stocktaking_items(q=None):
    """Return drug inventory rows used by printable stocktaking exports."""
    query = Item.query.filter_by(is_drug=True)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Item.name.ilike(like), Item.sku.ilike(like)))
    return query.order_by(Item.name.asc()).all()


def _stocktaking_filename(extension):
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    return f"evana_paed_stocktaking_{stamp}.{extension}"

bp = Blueprint("inventory", __name__)

def normalize_payer_name(name: str) -> str:
    """Normalize UI/query payer names so they match DB values more reliably."""
    if not name:
        return "Cash"
    n = (name or "").strip().lower().replace("insurance", "").strip()
    mapping = {
        "aar": "AAR",
        "apa": "APA",
        "cic": "CIC",
        "ga": "GA",
        "icea": "ICEA",
        "prudential": "Prudential",
        "sanlam": "Sanlam",
        "medicard": "Medicard",
        "case": "Case",
        "cash": "Cash",
    }
    return mapping.get(n, name.strip())


# ---------------------------
# Items (create/edit + quick price save)
# ---------------------------
@bp.route("/inventory/items", methods=["GET"])
@login_required
@roles_required("admin", "nurse")
def items_prices():
    """Back-compat endpoint (kept). Now shows Inventory/PriceBook/Merged via inventory_stock.html."""
    # Delegate to the main inventory page, default to inventory mode
    return redirect(url_for("inventory.inventory_stock", mode=request.args.get("mode") or "inventory"))

@bp.route("/inventory/items/save", methods=["POST"])
@login_required
@roles_required("admin")
def items_prices_save():
    items = Item.query.all()
    changed = 0
    for it in items:
        field = f"price_{it.id}"
        if field in request.form:
            try:
                new_price = Decimal(request.form.get(field) or "0")
            except Exception:
                new_price = None
            if new_price is not None and (it.sell_price or Decimal("0")) != new_price:
                it.sell_price = new_price
                changed += 1
    db.session.commit()
    flash(f"Item prices saved. {changed} change(s).", "success")
    return redirect(url_for("inventory.inventory_stock"))


@bp.route("/inventory/pricebook/<int:book_id>/save", methods=["POST"])
@login_required
@roles_required("admin")
def pricebook_prices_save(book_id):
    """Allow admins to update sell/buy prices for a selected price book."""
    book = PriceBook.query.get_or_404(book_id)
    changed = 0

    rows = PriceItem.query.filter(PriceItem.pricebook_id == book.id).all()
    for row in rows:
        sell_key = f"sell_{row.id}"
        buy_key = f"buy_{row.id}"

        if sell_key in request.form:
            raw_sell = (request.form.get(sell_key) or "").strip()
            if raw_sell:
                try:
                    sell_price = Decimal(raw_sell)
                except Exception:
                    flash(f"Invalid sell price for {row.item_name}.", "warning")
                    return redirect(
                        url_for(
                            "inventory.inventory_stock",
                            mode="pricebook",
                            payer=book.payer.name,
                            book_id=book.id,
                        )
                    )
                if row.sell_price != sell_price:
                    row.sell_price = sell_price
                    changed += 1

        if buy_key in request.form:
            raw_buy = (request.form.get(buy_key) or "").strip()
            if not raw_buy:
                if row.buy_price is not None:
                    row.buy_price = None
                    changed += 1
            else:
                try:
                    buy_price = Decimal(raw_buy)
                except Exception:
                    flash(f"Invalid buy price for {row.item_name}.", "warning")
                    return redirect(
                        url_for(
                            "inventory.inventory_stock",
                            mode="pricebook",
                            payer=book.payer.name,
                            book_id=book.id,
                        )
                    )
                if row.buy_price != buy_price:
                    row.buy_price = buy_price
                    changed += 1

    db.session.commit()
    flash(f"Updated {changed} price value(s) for {book.name}.", "success")
    return redirect(
        url_for(
            "inventory.inventory_stock",
            mode="pricebook",
            payer=book.payer.name,
            book_id=book.id,
        )
    )

@bp.route("/inventory/items/new", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def items_new():
    form = ItemPriceForm()
    if form.validate_on_submit():
        it = Item(
            name=form.name.data.strip(),
            sku=(form.sku.data or "").strip() or None,
            unit=(form.unit.data or "unit").strip(),
            min_level=0,
            current_qty=0,
            is_drug=bool(form.is_drug.data),
            sell_price=form.sell_price.data or 0
        )
        db.session.add(it)
        db.session.commit()
        flash("Item created.", "success")
        return redirect(url_for("inventory.inventory_stock", mode="inventory"))
    elif request.method == "POST":
        flash("Please correct the errors below.", "danger")
    return render_template("items_new.html", form=form)

@bp.route("/inventory/items/<int:item_id>/edit", methods=["GET","POST"])
@login_required
@roles_required("admin")
def items_edit(item_id):
    it = Item.query.get_or_404(item_id)
    form = ItemPriceForm(obj=it)
    if form.validate_on_submit():
        it.name = form.name.data.strip()
        it.sku = (form.sku.data or "").strip() or None
        it.unit = (form.unit.data or "unit").strip()
        it.is_drug = bool(form.is_drug.data)
        it.sell_price = form.sell_price.data or 0
        db.session.commit()
        flash("Item updated.", "success")
        return redirect(url_for("inventory.inventory_stock", mode="inventory"))
    # ensure initial values
    form.name.data = it.name
    form.sku.data = it.sku
    form.unit.data = it.unit or "unit"
    form.sell_price.data = it.sell_price or 0
    form.is_drug.data = it.is_drug
    return render_template("items_new.html", form=form)

# ---------------------------
# Inventory / Price Books / Merged
# ---------------------------
def _list_price_books(limit=50):
    q = (
        db.session.query(PriceBook, Payer.name.label("payer_name"))
        .join(Payer, PriceBook.payer_id == Payer.id)
        .order_by(PriceBook.effective_date.desc().nullslast(), PriceBook.id.desc())
        .limit(limit)
    )
    return q.all()

@bp.route("/inventory/stock")
@login_required
@roles_required("admin", "nurse")
def inventory_stock():
    """
    mode:
      - inventory : show Item table
      - pricebook : show PriceItem rows from selected/last book
      - merged    : Item LEFT JOIN selected PriceBook by SKU/Name
    query params: mode, payer, book_id, q
    """
    mode = (request.args.get("mode") or "inventory").strip().lower()
    payer_name_raw = (request.args.get("payer") or "Cash").strip()
    payer_name = normalize_payer_name(payer_name_raw)
    book_id = request.args.get("book_id", type=int)
    q = (request.args.get("q") or "").strip()

    # Resolve book for pricebook/merged modes
    book = None
    if mode in ("pricebook", "merged"):
        requested_payer = payer_name
        if book_id:
            # Keep requested book only when it belongs to the requested payer.
            candidate = db.session.query(PriceBook).filter(PriceBook.id == book_id).first()
            if candidate:
                payer_row = db.session.query(Payer).filter(Payer.id == candidate.payer_id).first()
                if payer_row and (payer_row.name or "").strip().lower() == requested_payer.strip().lower():
                    book = candidate
                    payer_name = payer_row.name

        if not book:
            # Fall back to "latest book for (case-insensitive, trimmed) payer name"
            qb = (
                db.session.query(PriceBook)
                .join(Payer, PriceBook.payer_id == Payer.id)
                .filter(func.lower(func.trim(Payer.name)) == requested_payer.strip().lower())
                .order_by(PriceBook.effective_date.desc().nullslast(), PriceBook.id.desc())
            )
            book = qb.first()
            if book:
                payer_row = db.session.query(Payer).filter(Payer.id == book.payer_id).first()
                if payer_row:
                    payer_name = payer_row.name
            else:
                flash(f"No price book found for payer '{requested_payer}'.", "warning")


    # Build data
    items = []
    rows_price = []
    rows_merged = []

    if mode == "inventory" or not book:
        query = Item.query
        if q:
            like = f"%{q}%"
            query = query.filter(or_(Item.name.ilike(like), Item.sku.ilike(like)))
        items = query.order_by(Item.name.asc()).all()

    if mode == "pricebook" and book:
        piq = PriceItem.query.filter(PriceItem.pricebook_id == book.id)
        if q:
            like = f"%{q}%"
            piq = piq.filter(or_(PriceItem.item_name.ilike(like), PriceItem.item_code.ilike(like)))
        rows_price = piq.order_by(PriceItem.item_type.asc(), PriceItem.item_name.asc()).all()

    if mode == "merged" and book:
        pi = db.aliased(PriceItem)
        sub = db.session.query(pi).filter(pi.pricebook_id == book.id).subquery()
        query = (
            db.session.query(
                Item.id.label("item_id"),
                Item.name.label("item_name"),
                Item.sku.label("item_sku"),
                Item.unit.label("item_unit"),
                Item.current_qty.label("qty"),
                Item.sell_price.label("inv_sell"),
                Item.buying_price.label("inv_buy"),
                sub.c.item_type.label("pb_type"),
                sub.c.sell_price.label("pb_sell"),
                sub.c.buy_price.label("pb_buy"),
                sub.c.unit.label("pb_unit"),
                sub.c.item_code.label("pb_code"),
                sub.c.item_name.label("pb_name"),
            )
            .outerjoin(
                sub,
                or_(
                    and_(sub.c.item_code.isnot(None), sub.c.item_code == Item.sku),
                    func.lower(sub.c.item_name) == func.lower(Item.name),
                ),
            )
        )
        if q:
            like = f"%{q}%"
            query = query.filter(
                or_(Item.name.ilike(like), Item.sku.ilike(like), sub.c.item_name.ilike(like))
            )
        rows_merged = query.order_by(Item.name.asc()).all()

    low = [it for it in items if (it.min_level or 0) >= (it.current_qty or 0)]
    return render_template(
        "inventory_stock.html",
        mode=mode,
        payer_name=payer_name,
        book=book,
        books=_list_price_books(),
        items=items,
        rows_price=rows_price,
        rows_merged=rows_merged,
        low_count=len(low),
        q=q
    )



@bp.route("/inventory/stocktaking/export")
@login_required
@roles_required("admin", "nurse")
def inventory_stocktaking_export():
    """Download a printable stocktaking form for pharmacy drug counts."""
    export_format = (request.args.get("format") or "pdf").strip().lower()
    q = (request.args.get("q") or "").strip()
    items = _stocktaking_items(q=q)

    if export_format in {"excel", "xlsx"}:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Stocktaking"
        ws.append(["Evana Paed Pharmacy Stocktaking Form"])
        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
        ws.append([])
        headers = [
            "No.",
            "Drug Name",
            "SKU",
            "Unit",
            "Quantity in Evana-Paed System",
            "Actual Quantity in Pharmacy Stock",
            "Variance",
            "Remarks",
        ]
        ws.append(headers)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for idx, item in enumerate(items, start=1):
            row_num = ws.max_row + 1
            ws.append([
                idx,
                item.name,
                item.sku or "",
                item.unit or "",
                int(item.current_qty or 0),
                "",
                f"=F{row_num}-E{row_num}",
                "",
            ])

        widths = [8, 34, 16, 12, 24, 28, 14, 24]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A5"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=_stocktaking_filename("xlsx"),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    if export_format != "pdf":
        flash("Unsupported stocktaking export format.", "warning")
        return redirect(url_for("inventory.inventory_stock", mode="inventory", q=q))

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except Exception:
        html_rows = "".join(
            f"<tr><td>{idx}</td><td>{item.name}</td><td>{item.sku or ''}</td><td>{item.unit or ''}</td>"
            f"<td>{int(item.current_qty or 0)}</td><td></td><td></td><td></td></tr>"
            for idx, item in enumerate(items, start=1)
        )
        html = f"""
        <h1>Evana Paed Pharmacy Stocktaking Form</h1>
        <p>Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}</p>
        <table border="1" cellspacing="0" cellpadding="4">
          <thead><tr><th>No.</th><th>Drug Name</th><th>SKU</th><th>Unit</th><th>Quantity in Evana-Paed System</th><th>Actual Quantity in Pharmacy Stock</th><th>Variance</th><th>Remarks</th></tr></thead>
          <tbody>{html_rows}</tbody>
        </table>
        """
        return Response(html, mimetype="text/html")

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Evana Paed Pharmacy Stocktaking Form", styles["Title"]),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [["No.", "Drug Name", "SKU", "Unit", "Quantity in Evana-Paed System", "Actual Quantity in Pharmacy Stock", "Variance", "Remarks"]]
    for idx, item in enumerate(items, start=1):
        data.append([idx, item.name, item.sku or "", item.unit or "", int(item.current_qty or 0), "", "", ""])
    table = Table(data, repeatRows=1, colWidths=[32, 170, 70, 55, 105, 125, 65, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={_stocktaking_filename('pdf')}"},
    )

# ---------------------------
# Stock transaction
# ---------------------------
@bp.route("/inventory/stock/txn", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def inventory_txn():
    form = StockTxnForm()
    items = Item.query.order_by(Item.name.asc()).all()
    form.item_id.choices = [(it.id, f"{it.name} ({it.unit or ''}) • In stock: {it.current_qty}") for it in items]
    if form.validate_on_submit():
        it = Item.query.get_or_404(form.item_id.data)
        try:
            qty_change_int = int(Decimal(str(form.qty.data)))
        except Exception:
            flash("Quantity must be a whole number.", "danger")
            return redirect(url_for("inventory.inventory_txn"))
        it.current_qty = int(it.current_qty or 0) + qty_change_int
        db.session.add(
            ItemTxn(
                item_id=it.id,
                qty_change=qty_change_int,
                reason=form.reason.data,
                note=(form.note.data or "").strip(),
            )
        )
        db.session.commit()
        flash(f"Stock updated for {it.name}. New qty: {it.current_qty}", "success")
        return redirect(url_for("inventory.inventory_stock"))
    elif request.method == "POST":
        flash("Please correct the errors below.", "danger")
    return render_template("inventory_txn.html", form=form)

@bp.route("/inventory/stock/txn/bulk", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def inventory_bulk_txn():
    form = BulkStockTxnForm()
    items = Item.query.filter_by(is_drug=True).order_by(Item.name.asc()).all()

    if form.validate_on_submit():
        changes = []
        for it in items:
            raw_qty = (request.form.get(f"qty_{it.id}") or "").strip()
            if not raw_qty:
                continue
            try:
                qty_change_int = int(Decimal(raw_qty))
            except Exception:
                flash(f"{it.name}: quantity must be a whole number.", "danger")
                return redirect(url_for("inventory.inventory_bulk_txn"))
            if qty_change_int == 0:
                continue
            changes.append((it, qty_change_int))

        if not changes:
            flash("No quantity changes entered.", "warning")
            return redirect(url_for("inventory.inventory_bulk_txn"))

        note = (form.note.data or "").strip()
        for it, qty_change_int in changes:
            it.current_qty = int(it.current_qty or 0) + qty_change_int
            db.session.add(
                ItemTxn(
                    item_id=it.id,
                    qty_change=qty_change_int,
                    reason=form.reason.data,
                    note=note,
                    user_id=current_user.id,
                )
            )

        db.session.commit()
        flash(f"Bulk stock transaction recorded for {len(changes)} drug(s).", "success")
        return redirect(url_for("inventory.inventory_stock", mode="inventory"))
    elif request.method == "POST":
        flash("Please correct the errors below.", "danger")

    return render_template("inventory_bulk_txn.html", form=form, items=items)

# ---------------------------
# Dispense (as you had)
# ---------------------------
@bp.route("/patients/<int:patient_id>/visits/<int:visit_id>/dispense", methods=["GET", "POST"])
@login_required
@roles_required("admin", "nurse")
def visit_dispense(patient_id, visit_id):
    p = Patient.query.get_or_404(patient_id)
    v = Visit.query.filter_by(id=visit_id, patient_id=patient_id).first_or_404()
    inv = Invoice.query.filter_by(visit_id=visit_id, patient_id=patient_id).first()
    rows = []
    if inv:
        disp_by_line = defaultdict(Decimal)
        disp_by_item = defaultdict(Decimal)
        for d in DispenseTxn.query.filter_by(visit_id=visit_id).all():
            if d.invoice_line_id:
                disp_by_line[d.invoice_line_id] += Decimal(d.qty or 0)
            elif d.item_id:
                disp_by_item[d.item_id] += Decimal(d.qty or 0)
        for ln in inv.lines:
            if not (ln.kind == "drug" or ln.item_id):
                continue
            item = Item.query.get(ln.item_id) if ln.item_id else None
            stock = item.current_qty if item else 0
            prescribed = Decimal(ln.qty or 0)
            dispensed = disp_by_line.get(ln.id, Decimal("0"))
            if dispensed == 0 and ln.item_id:
                dispensed = disp_by_item.get(ln.item_id, Decimal("0"))
            remaining = max(Decimal("0"), prescribed - dispensed)
            rows.append(
                {
                    "line_id": ln.id,
                    "item": item,
                    "item_id": ln.item_id,
                    "desc": ln.description,
                    "unit_price": Decimal(ln.unit_price or 0),
                    "prescribed_qty": prescribed,
                    "dispensed_qty": dispensed,
                    "remaining_qty": remaining,
                    "stock": stock,
                }
            )
    if request.method == "POST":
        line_ids = request.form.getlist("line_id[]")
        item_ids = request.form.getlist("item_id[]")
        qtys = request.form.getlist("qty[]")
        if not line_ids:
            flash("No items to dispense.", "warning")
            return redirect(url_for("inventory.visit_dispense", patient_id=patient_id, visit_id=visit_id))
        remaining_map = {str(r["line_id"]): r["remaining_qty"] for r in rows}
        dispensed_any, errors = False, []
        for lid, iid, qraw in zip(line_ids, item_ids, qtys):
            try:
                q = Decimal((qraw or "0")).quantize(Decimal("1"))
            except Exception:
                q = Decimal("0")
            if q <= 0:
                continue
            remaining = remaining_map.get(lid, Decimal("0"))
            if q > remaining:
                errors.append(f"Line #{lid}: quantity {q} exceeds remaining {remaining}.")
                continue
            it = Item.query.get(int(iid)) if iid and iid.isdigit() else None
            if not it:
                errors.append(f"Line #{lid}: item not found.")
                continue
            current_stock = int(it.current_qty or 0)
            if q > current_stock:
                errors.append(f"{it.name}: only {current_stock} in stock, cannot dispense {q}.")
                continue
            it.current_qty = current_stock - int(q)
            db.session.add(
                DispenseTxn(
                    item_id=it.id,
                    patient_id=patient_id,
                    visit_id=v.id,
                    invoice_line_id=int(lid) if lid and lid.isdigit() else None,
                    qty=q,
                    unit_price=Decimal(it.sell_price or 0),
                    line_total=Decimal(it.sell_price or 0) * q,
                )
            )
            db.session.add(
                ItemTxn(
                    item_id=it.id,
                    qty_change=int(-q),
                    reason="Consume-Visit",
                    visit_id=v.id,
                    note=f"Dispensed for visit {v.id}",
                )
            )
            dispensed_any = True
        if errors:
            for e in errors:
                flash(e, "warning")
        if dispensed_any:
            db.session.commit()
            flash("Dispense recorded and inventory updated.", "success")
            return redirect(url_for("patients.patient_chart", patient_id=patient_id, tab="billing"))
        return redirect(url_for("inventory.visit_dispense", patient_id=patient_id, visit_id=visit_id))
    rows_to_show = [r for r in rows if r["remaining_qty"] > 0]
    return render_template("visit_dispense.html", p=p, v=v, rows=rows_to_show)
